import streamlit as st
import pandas as pd
import os
from pathlib import Path
import tempfile
from openpyxl import load_workbook
import zipfile
import io


def process_home_visit_data_by_field_mapping(data_file, output_dir):
    """
    处理家访表数据：在一个工作簿中创建多个工作表，每个工作表对应一个学生
    只处理"是否提交"为"是"的数据行

    :param data_file: 包含数据的Excel文件路径
    :param output_dir: 输出目录路径
    :return: 生成的文件列表
    """
    # 读取数据文件
    data_df = pd.read_excel(data_file)

    # 固定模板文件路径
    template_file = "/workspaces/tools/template/家访表样表.xlsx"

    # 检查模板文件是否存在
    if not os.path.exists(template_file):
        raise FileNotFoundError(f"模板文件不存在: {template_file}")

    # 过滤只处理"是否提交"为"是"的数据行
    if "是否提交" in data_df.columns:
        filtered_data_df = data_df[data_df["是否提交"] == "是"].copy()
        st.info(f"共找到 {len(filtered_data_df)} 条'是否提交'为'是'的记录")
    else:
        st.warning("警告：数据中没有找到'是否提交'列，将处理所有记录")
        filtered_data_df = data_df

    generated_files = []

    # 加载模板工作簿
    workbook = load_workbook(template_file)

    # 获取模板工作表作为基准
    template_sheet = workbook.active
    template_sheet.title = "模板"  # 将原始模板重命名，避免名称冲突

    # 遍历过滤后的数据中的每一行
    for idx, row in filtered_data_df.iterrows():
        # 复制模板工作表（第一条数据直接在模板上修改，后续数据复制新工作表）
        if idx == 0:
            # 第一条数据使用模板工作表
            worksheet = template_sheet
            worksheet.title = "第一个学生"  # 给第一个学生一个占位名
        else:
            # 复制模板工作表
            worksheet = workbook.copy_worksheet(template_sheet)

        # 获取学生姓名作为工作表名称
        student_name_col = (
            data_df.columns[0] if len(data_df.columns) > 0 else "学生姓名"
        )
        student_name = (
            str(row[student_name_col])
            if pd.notna(row[student_name_col])
            else f"学生_{idx+1}"
        )

        # 清理工作表名称（Excel工作表名称有限制：最多31个字符，不能包含: \ / ? * [ ]）
        sheet_name = str(student_name)[:31]  # 截断到31个字符
        # 移除Excel不允许的特殊字符
        invalid_chars = [":", "\\", "/", "?", "*", "[", "]"]
        for char in invalid_chars:
            sheet_name = sheet_name.replace(char, "_")

        # 确保工作表名称不为空
        if not sheet_name.strip():
            sheet_name = f"学生_{idx+1}"

        # 重命名工作表
        worksheet.title = sheet_name

        # 遍历数据的每一列，将值填入模板中对应的右边单元格
        for col_name, value in row.items():
            # 跳过"是否提交"列本身，因为这只是筛选条件
            if col_name == "是否提交":
                continue

            # 尝试找到模板中匹配的列名
            found_match = False

            # 搜索整个工作表查找匹配的单元格内容
            for row_search in range(1, worksheet.max_row + 1):
                for col_search in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_search, column=col_search)
                    if cell.value and str(cell.value).strip() == str(col_name).strip():
                        # 找到匹配的字段名，将值填入右边的单元格
                        worksheet.cell(
                            row=row_search,
                            column=col_search + 1,
                            value=value if pd.notna(value) else "",
                        )
                        found_match = True
                        break
                if found_match:
                    break

        # 如果需要，可以在这里添加特殊处理：比如更新每个工作表中的某些特定字段

        # 更新工作表名称（以防需要在填充数据后更新名称）
        # 这里不需要再次设置，已经在上面设置了

    # 保存工作簿（此时所有学生的数据都在同一个工作簿的不同工作表中）
    output_file = os.path.join(output_dir, "家访表_所有学生.xlsx")
    workbook.save(output_file)
    generated_files.append(output_file)

    return generated_files


def show_home_visit_page():
    """
    Streamlit界面函数，用于展示家访表数据填充功能
    """
    st.title("🏠 家访表数据填充工具")

    # 提供模板文件说明
    st.info(
        "此工具将把家访表数据中的每条记录添加到同一个工作簿的不同工作表中，以学生姓名命名工作表。"
    )
    st.info("注意：系统将自动使用template目录下的家访表样表.xlsx作为模板，无需上传。")

    # 侧边栏 - 文件上传和操作功能
    with st.sidebar:
        st.header("📋 家访表管理")

        # 选择数据文件（只上传数据文件）
        data_file = st.file_uploader(
            "上传家访表数据文件 (Excel)", type=["xlsx"], key="data"
        )

        # 检查template目录中的模板文件
        template_file = "/workspaces/tools/template/家访表样表.xlsx"
        if not os.path.exists(template_file):
            st.error(f"错误：找不到模板文件 {template_file}")
            return
        else:
            st.success(f"✅ 已找到模板文件: 家访表样表.xlsx")

        # 如果用户没有上传数据文件，使用默认数据文件
        data_path = None

        if data_file is not None:
            # 将上传的文件保存到临时位置
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
                tmp_file.write(data_file.getvalue())
                data_path = tmp_file.name
        else:
            # 使用默认数据文件
            default_data_path = "/workspaces/tools/template/家访表数据.xlsx"
            if os.path.exists(default_data_path):
                data_path = default_data_path
                st.info(f"使用默认数据文件: 家访表数据.xlsx")

        # 开始处理按钮
        generate_clicked = st.button("开始生成家访表", type="primary")

        # 初始化会话状态
        if "generated_files_zip" not in st.session_state:
            st.session_state.generated_files_zip = None
        if "generated_count" not in st.session_state:
            st.session_state.generated_count = 0

        # 处理生成请求
        if generate_clicked:
            if data_path:
                try:
                    with st.spinner("正在生成家访表，请稍候..."):
                        # 创建临时输出目录
                        with tempfile.TemporaryDirectory() as temp_dir:
                            # 处理数据，将所有学生放在同一个工作簿的不同工作表中
                            generated_files = process_home_visit_data_by_field_mapping(
                                data_path, temp_dir
                            )

                            if len(generated_files) == 0:
                                st.warning(
                                    "⚠️ 没有找到'是否提交'为'是'的记录，未生成任何文件"
                                )
                            else:
                                # 显示生成的文件信息
                                st.success(
                                    f"✅ 已成功生成 1 个工作簿文件，包含多个工作表！"
                                )
                                st.info(
                                    f"文件名: {os.path.basename(generated_files[0])}"
                                )

                                # 将生成的文件内容提供给下载
                                with open(generated_files[0], "rb") as f:
                                    file_content = f.read()

                                # 保存文件数据到会话状态
                                st.session_state.generated_files_zip = file_content
                                st.session_state.generated_count = 1

                except Exception as e:
                    st.error(f"❌ 发生错误: {str(e)}")
                    import traceback

                    st.code(traceback.format_exc())

                finally:
                    # 清理临时文件
                    if data_file is not None and data_path:
                        os.unlink(data_path)
            else:
                st.warning("请上传数据文件，或确保系统中有默认数据文件。")

        # 如果有生成的文件，显示下载按钮
        if st.session_state.generated_files_zip:
            # 提供一个选择：下载单个文件还是ZIP（虽然只有一个文件）
            col1, col2 = st.columns(2)
            with col1:
                st.download_button(
                    label="📥 下载家访表工作簿",
                    data=st.session_state.generated_files_zip,
                    file_name="家访表_所有学生.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            with col2:
                # 也可以打包成ZIP（虽然只有一个文件）
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
                    zip_file.writestr(
                        "家访表_所有学生.xlsx", st.session_state.generated_files_zip
                    )

                st.download_button(
                    label="📦 下载为ZIP",
                    data=zip_buffer.getvalue(),
                    file_name="家访表.zip",
                    mime="application/zip",
                )

        # 在侧边栏最下方添加模板下载功能
        st.markdown("---")
        st.subheader("📥 模板下载")

        # 下载数据模板
        with open("/workspaces/tools/template/家访表数据.xlsx", "rb") as file:
            st.download_button(
                label="家访表数据模板",
                data=file,
                file_name="家访表数据.xlsx",
                mime="ms-excel",
            )

        # 下载样表模板
        with open("/workspaces/tools/template/家访表样表.xlsx", "rb") as file:
            st.download_button(
                label="家访表样表模板",
                data=file,
                file_name="家访表样表.xlsx",
                mime="ms-excel",
            )

        st.markdown("---")

        st.warning(
            "💡 **重要提示：**\n"
            "1. 上传家访表数据文件，系统将自动使用家访表样表作为模板\n"
            "2. 只有'是否提交'为'是'的数据才会被处理\n"
            "3. **新功能**：所有学生将放在同一个工作簿的不同工作表中"
        )

    # 主内容区域 - 显示数据预览
    if data_file is not None:
        # 显示数据文件的预览
        data_preview = pd.read_excel(data_file)
        st.info(f"数据文件预览 ({len(data_preview)} 行)")
        st.dataframe(data_preview)
    else:
        # 显示默认数据文件的预览
        default_data_path = "/workspaces/tools/template/家访表数据.xlsx"
        if os.path.exists(default_data_path):
            data_preview = pd.read_excel(default_data_path)
            st.info(f"默认数据文件预览 ({len(data_preview)} 行)")
            st.dataframe(data_preview)
        else:
            st.info("请上传家访表数据文件")

    # 添加一个关于新功能的说明
    with st.expander("ℹ️ 使用说明"):
        st.markdown(
            """
        **如何使用家访表批量生成工具：**
        
        1. 在侧边栏上传包含家访数据的Excel文件（每行代表一个学生的家访信息）
        2. 系统将自动使用template目录下的家访表样表.xlsx作为模板
        3. 点击"开始生成家访表"按钮
        4. 在侧边栏下载生成的Excel工作簿
        
        **新功能特点：**
        - **所有学生在同一个工作簿**：不再为每个学生生成单独的文件
        - **每个学生一个工作表**：工作簿中的每个工作表对应一个学生
        - **工作表以学生姓名命名**：便于查找和浏览
        - **保留模板格式**：每个工作表都使用模板的格式和样式
        
        **功能细节：**
        - 系统只会处理'是否提交'列为'是'的记录
        - 数据会根据列标题自动映射到样表中字段名右边的单元格
        - 保留样表的格式和样式，仅替换数据部分
        - 如果未上传数据文件，系统将使用默认数据文件进行演示
        
        **注意事项：**
        - 数据文件中的列名应与样表中的字段名匹配
        - 数据将填充到匹配字段名右边的单元格
        - 工作表名称有31个字符限制，长姓名会被截断
        - 工作表名称不能包含特殊字符: \\ / ? * [ ]
        - 支持的文件格式：`.xlsx`
        
        **文件结构示例：**
        ```
        家访表_所有学生.xlsx
        ├── 工作表1: 张三
        ├── 工作表2: 李四
        ├── 工作表3: 王五
        └── 工作表4: 模板 (原始模板备份)
        ```
        """
        )


if __name__ == "__main__":
    # 设置页面配置
    st.set_page_config(
        page_title="家访表数据填充工具 - 单工作簿版", page_icon="🏠", layout="wide"
    )

    # 运行界面函数
    show_home_visit_page()
