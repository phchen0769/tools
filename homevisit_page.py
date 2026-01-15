import streamlit as st
import pandas as pd
import os
from pathlib import Path
import tempfile
from openpyxl import load_workbook
import zipfile
import io


def generate_unique_sheet_name(workbook, base_name):
    """
    生成唯一的工作表名称，避免重复
    :param workbook: 工作簿对象
    :param base_name: 基础名称（通常是学生姓名）
    :return: 唯一的工作表名称
    """
    # 获取所有现有工作表名称
    existing_names = [sheet.title for sheet in workbook.worksheets]

    # 如果基础名称不存在，直接返回
    if base_name not in existing_names:
        return base_name

    # 如果存在，添加数字后缀直到找到不重复的名称
    counter = 1
    while True:
        new_name = f"{base_name}_{counter}"
        if new_name not in existing_names:
            return new_name
        counter += 1


def process_home_visit_data_by_field_mapping(data_file, output_dir):
    """
    处理家访表数据：在一个工作簿中创建多个工作表，每个工作表对应一个学生
    只处理"是否提交"为"是"的数据行

    :param data_file: 包含数据的Excel文件路径
    :param output_dir: 输出目录路径
    :return: 生成的文件列表
    """
    # 读取数据文件
    data_df = pd.read_excel(data_file)

    # 固定模板文件路径
    template_file = "/workspaces/tools/template/家访表样表.xlsx"

    # 检查模板文件是否存在
    if not os.path.exists(template_file):
        raise FileNotFoundError(f"模板文件不存在: {template_file}")

    # 过滤只处理"是否提交"为"是"的数据行
    if "是否提交" in data_df.columns:
        filtered_data_df = data_df[data_df["是否提交"] == "是"].copy()
        st.info(f"共找到 {len(filtered_data_df)} 条'是否提交'为'是'的记录")
    else:
        st.warning("警告：数据中没有找到'是否提交'列，将处理所有记录")
        filtered_data_df = data_df

    generated_files = []

    # 加载模板工作簿
    workbook = load_workbook(template_file)

    # 获取模板工作表作为基准
    template_sheet = workbook.active

    # 遍历过滤后的数据中的每一行
    for idx, row in filtered_data_df.iterrows():
        # 如果是第一条数据，直接在模板工作表上修改
        if idx == 0:
            worksheet = template_sheet
        else:
            # 复制模板工作表为新工作表
            worksheet = workbook.copy_worksheet(template_sheet)

        # 获取学生姓名 - 优先查找"学生姓名"列，如果没有则用第一列
        student_name = ""

        # 尝试从常见的姓名列名中查找
        possible_name_columns = ["学生姓名", "姓名", "name", "姓名", "学生名称"]
        for col_name in possible_name_columns:
            if col_name in row and pd.notna(row[col_name]):
                student_name = str(row[col_name]).strip()
                break

        # 如果没找到，使用第一列
        if not student_name and len(data_df.columns) > 0:
            first_col = data_df.columns[0]
            if first_col in row and pd.notna(row[first_col]):
                student_name = str(row[first_col]).strip()

        # 如果还是没有找到姓名，使用索引
        if not student_name:
            student_name = f"学生_{idx+1}"

        # 清理工作表名称（Excel工作表名称有限制：最多31个字符，不能包含: \ / ? * [ ]）
        sheet_name = str(student_name)[:31]  # 截断到31个字符
        # 移除Excel不允许的特殊字符
        invalid_chars = [":", "\\", "/", "?", "*", "[", "]"]
        for char in invalid_chars:
            sheet_name = sheet_name.replace(char, "_")

        # 移除开头的空格和点号
        sheet_name = sheet_name.strip(". ")

        # 确保工作表名称不为空
        if not sheet_name.strip():
            sheet_name = f"学生_{idx+1}"

        # 生成唯一的工作表名称（避免重复）
        unique_sheet_name = generate_unique_sheet_name(workbook, sheet_name)

        # 重命名工作表
        worksheet.title = unique_sheet_name

        # 遍历数据的每一列，将值填入模板中对应的右边单元格
        for col_name, value in row.items():
            # 跳过"是否提交"列本身，因为这只是筛选条件
            if col_name == "是否提交":
                continue

            # 特殊处理联系电话字段，在值的开头增加一个单引号
            cell_value = value if pd.notna(value) else ""
            if col_name == "联系电话" and pd.notna(value):
                cell_value = "'" + str(value)

            # 尝试找到模板中匹配的列名
            found_match = False

            # 搜索整个工作表查找匹配的单元格内容
            for row_search in range(1, worksheet.max_row + 1):
                for col_search in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_search, column=col_search)
                    if cell.value and str(cell.value).strip() == str(col_name).strip():
                        # 找到匹配的字段名，将值填入右边的单元格
                        worksheet.cell(
                            row=row_search, column=col_search + 1, value=cell_value
                        )
                        found_match = True
                        break
                if found_match:
                    break

    # 删除原始的模板工作表（如果有的话）
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    # 保存工作簿（此时所有学生的数据都在同一个工作簿的不同工作表中）
    output_file = os.path.join(output_dir, "家访表_所有学生.xlsx")
    workbook.save(output_file)
    generated_files.append(output_file)

    return generated_files


def show_home_visit_page():
    """
    Streamlit界面函数，用于展示家访表数据填充功能
    """
    # st.title("🏠 家访表数据填充工具")

    # 提供模板文件说明
    st.info(
        "此工具将把家访表数据中的每条记录添加到同一个工作簿的不同工作表中，每个工作表以对应学生姓名命名。"
    )
    st.info("注意：系统将自动使用template目录下的家访表样表.xlsx作为模板，无需上传。")

    # 侧边栏 - 文件上传和操作功能
    with st.sidebar:
        st.header("📋 家访表管理")

        # 选择数据文件（只上传数据文件）
        data_file = st.file_uploader(
            "上传家访表数据文件 (Excel)", type=["xlsx"], key="data"
        )

        # 检查template目录中的模板文件
        template_file = "/workspaces/tools/template/家访表样表.xlsx"
        if not os.path.exists(template_file):
            st.error(f"错误：找不到模板文件 {template_file}")
            return
        else:
            st.success(f"✅ 已找到模板文件: 家访表样表.xlsx")

        # 如果用户没有上传数据文件，使用默认数据文件
        data_path = None

        if data_file is not None:
            # 将上传的文件保存到临时位置
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
                tmp_file.write(data_file.getvalue())
                data_path = tmp_file.name
        else:
            # 使用默认数据文件
            default_data_path = "/workspaces/tools/template/家访表数据.xlsx"
            if os.path.exists(default_data_path):
                data_path = default_data_path
                st.info(f"使用默认数据文件: 家访表数据.xlsx")

        # 开始处理按钮
        generate_clicked = st.button("开始生成家访表", type="primary")

        # 初始化会话状态
        if "generated_files_zip" not in st.session_state:
            st.session_state.generated_files_zip = None
        if "generated_count" not in st.session_state:
            st.session_state.generated_count = 0
        if "generated_sheet_names" not in st.session_state:
            st.session_state.generated_sheet_names = []

        # 处理生成请求
        if generate_clicked:
            if data_path:
                try:
                    with st.spinner("正在生成家访表，请稍候..."):
                        # 创建临时输出目录
                        with tempfile.TemporaryDirectory() as temp_dir:
                            # 处理数据，将所有学生放在同一个工作簿的不同工作表中
                            generated_files = process_home_visit_data_by_field_mapping(
                                data_path, temp_dir
                            )

                            if len(generated_files) == 0:
                                st.warning(
                                    "⚠️ 没有找到'是否提交'为'是'的记录，未生成任何文件"
                                )
                            else:
                                # 加载生成的工作簿，获取工作表名称
                                workbook = load_workbook(generated_files[0])
                                sheet_names = workbook.sheetnames

                                # 显示生成的文件信息
                                st.success(
                                    f"✅ 已成功生成 1 个工作簿文件，包含 {len(sheet_names)} 个工作表！"
                                )
                                st.info(
                                    f"文件名: {os.path.basename(generated_files[0])}"
                                )

                                # 显示工作表名称列表
                                if sheet_names:
                                    st.info(
                                        f"工作表名称: {', '.join(sheet_names[:10])}{'...' if len(sheet_names) > 10 else ''}"
                                    )

                                # 将生成的文件内容提供给下载
                                with open(generated_files[0], "rb") as f:
                                    file_content = f.read()

                                # 保存文件数据到会话状态
                                st.session_state.generated_files_zip = file_content
                                st.session_state.generated_count = len(sheet_names)
                                st.session_state.generated_sheet_names = sheet_names

                except Exception as e:
                    st.error(f"❌ 发生错误: {str(e)}")
                    import traceback

                    st.code(traceback.format_exc())

                finally:
                    # 清理临时文件
                    if data_file is not None and data_path:
                        os.unlink(data_path)
            else:
                st.warning("请上传数据文件，或确保系统中有默认数据文件。")

        # 如果有生成的文件，显示下载按钮
        if st.session_state.generated_files_zip:
            # 显示生成的工作表数量
            if st.session_state.generated_count > 0:
                st.success(f"📊 已生成 {st.session_state.generated_count} 个工作表")

            # 直接提供下载生成的Excel工作簿
            st.download_button(
                label="📥 下载家访表工作簿",
                data=st.session_state.generated_files_zip,
                file_name="家访表_所有学生.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        # 在侧边栏最下方添加模板下载功能
        st.markdown("---")
        st.subheader("📥 模板下载")

        # 下载数据模板
        with open("/workspaces/tools/template/家访表数据.xlsx", "rb") as file:
            st.download_button(
                label="家访表数据模板",
                data=file,
                file_name="家访表数据.xlsx",
                mime="ms-excel",
            )

        # 下载样表模板
        with open("/workspaces/tools/template/家访表样表.xlsx", "rb") as file:
            st.download_button(
                label="家访表样表模板",
                data=file,
                file_name="家访表样表.xlsx",
                mime="ms-excel",
            )

        st.markdown("---")

        st.warning(
            "💡 **重要提示：**\n"
            "1. 上传家访表数据文件，系统将自动使用家访表样表作为模板\n"
            "2. 只有'是否提交'为'是'的数据才会被处理\n"
            "3. **新功能**：所有学生将放在同一个工作簿的不同工作表中\n"
            "4. **工作表命名**：每个工作表将以学生姓名命名"
        )

    # 主内容区域 - 显示数据预览
    if data_file is not None:
        # 显示数据文件的预览
        data_preview = pd.read_excel(data_file)
        st.info(f"数据文件预览 ({len(data_preview)} 行)")
        st.dataframe(data_preview)

        # 检查是否有学生姓名列
        possible_name_columns = ["学生姓名", "姓名", "name", "姓名", "学生名称"]
        found_name_columns = [
            col for col in possible_name_columns if col in data_preview.columns
        ]
        if found_name_columns:
            st.success(f"✅ 检测到姓名列: {', '.join(found_name_columns)}")
        else:
            st.warning("⚠️ 未检测到常见的姓名列，将使用第一列数据作为学生姓名")
    else:
        # 显示默认数据文件的预览
        default_data_path = "/workspaces/tools/template/家访表数据.xlsx"
        if os.path.exists(default_data_path):
            data_preview = pd.read_excel(default_data_path)
            st.info(f"默认数据文件预览 ({len(data_preview)} 行)")
            st.dataframe(data_preview)
        else:
            st.info("请上传家访表数据文件")

    # 添加一个关于新功能的说明
    with st.expander("ℹ️ 使用说明"):
        st.markdown(
            """
        **如何使用家访表批量生成工具：**
        
        1. 在侧边栏上传包含家访数据的Excel文件（每行代表一个学生的家访信息）
        2. 系统将自动使用template目录下的家访表样表.xlsx作为模板
        3. 点击"开始生成家访表"按钮
        4. 在侧边栏下载生成的Excel工作簿
        
        **功能特点：**
        - **所有学生在同一个工作簿**：不再为每个学生生成单独的文件
        - **每个学生一个工作表**：工作簿中的每个工作表对应一个学生
        - **工作表以学生姓名命名**：便于查找和浏览
        - **智能姓名识别**：自动识别常见的姓名列
        - **唯一名称处理**：自动处理重复的学生姓名
        - **保留模板格式**：每个工作表都使用模板的格式和样式
        
        **功能细节：**
        - 系统只会处理'是否提交'列为'是'的记录
        - 数据会根据列标题自动映射到样表中字段名右边的单元格
        - 保留样表的格式和样式，仅替换数据部分
        - 如果未上传数据文件，系统将使用默认数据文件进行演示
        
        **注意事项：**
        - 数据文件中的列名应与样表中的字段名匹配
        - 数据将填充到匹配字段名右边的单元格
        - 工作表名称有31个字符限制，长姓名会被截断
        - 工作表名称不能包含特殊字符: \\ / ? * [ ]
        - 如果有重名的学生，系统会自动添加数字后缀（例如：张三_1, 张三_2）
        - 支持的文件格式：`.xlsx`
        """
        )


if __name__ == "__main__":
    # 设置页面配置
    st.set_page_config(
        # page_title="家访表数据填充工具 - 单工作簿版", page_icon="🏠", layout="wide"
    )

    # 运行界面函数
    show_home_visit_page()
